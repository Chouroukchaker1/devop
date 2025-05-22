import os
import xml.etree.ElementTree as ET
import pandas as pd
from datetime import datetime
import json
import sys
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

# Forcer l'encodage en UTF-8 pour éviter les erreurs d'affichage
sys.stdout.reconfigure(encoding='utf-8')

def format_date(date):
    if isinstance(date, str) and len(date) == 10:
        try:
            dt = datetime.strptime(date, "%d/%m/%Y")
            return dt.strftime('%d/%m/%Y')
        except ValueError:
            print(f"Erreur de format de date : {date}")
    return date

def extract_fuel_data(xml_string, directory):
    try:
        root = ET.fromstring(xml_string)
    except ET.ParseError as e:
        print(f"Erreur de parsing XML : {e}")
        return {}

    namespaces = {'ns': 'http://aeec.aviation-ia.net/633'}
    fuel_data = {"Directory": directory}

    # Extraction of fuel-related data, including 'FinalReserve' and 'Tankering Weight' (used for 'Additional Fuel (tonnes)')
    fuel_elements = {
        'TaxiFuel': './/ns:TaxiFuel/ns:EstimatedWeight/ns:Value',
        'TripFuel': './/ns:TripFuel/ns:EstimatedWeight/ns:Value',
        'ContingencyFuel': './/ns:ContingencyFuel/ns:EstimatedWeight/ns:Value',
        'BlockFuel': './/ns:BlockFuel/ns:EstimatedWeight/ns:Value',
        'FinalReserve': './/ns:FinalReserve/ns:EstimatedWeight/ns:Value',  # Added FinalReserve
        'Tankering Weight': './/ns:TankeringWeight/ns:Value'
    }

    for key, path in fuel_elements.items():
        element = root.find(path, namespaces)
        fuel_data[key] = float(element.text) / 1000 if element is not None and element.text.replace('.', '', 1).isdigit() else 0.0

    # Computation of 'Additional Fuel (tonnes)'
    tankering_advice = root.find('.//ns:TankeringAdvice', namespaces)
    is_tankering_sector = False

    if tankering_advice is not None and tankering_advice.text:
        tankering_text = tankering_advice.text.strip()
        is_tankering_sector = tankering_text.lower() == 'tankering sector'

    fuel_data['Additional Fuel (tonnes)'] = fuel_data['Tankering Weight'] if is_tankering_sector else 0.0

    # Extraction of 'Alternate Arrival Airport'
    alternate_airport = None
    for airport_data in root.findall('.//ns:AirportData', namespaces):
        airport = airport_data.find('ns:Airport', namespaces)
        if airport is not None and 'PrimaryArrivalAlternateAirport' in airport.get('airportFunction', ''):
            icao = airport_data.find('ns:AirportICAOCode', namespaces)
            if icao is not None and icao.text:
                fuel_data['Alternate Arrival Airport'] = icao.text
                alternate_airport = icao.text
                break
                
            icao = airport_data.find('ns:AirportICAOCode', namespaces)
            if icao is not None and icao.text:
                fuel_data['Alternate Arrival Airport'] = icao.text
                alternate_airport = icao.text
                break
                
            iata = airport_data.find('ns:AirportIATACode', namespaces)
            if iata is not None and iata.text:
                fuel_data['Alternate Arrival Airport'] = iata.text
                alternate_airport = iata.text
                break
                
            iata = airport.find('ns:AirportIATACode', namespaces)
            if iata is not None and iata.text:
                fuel_data['Alternate Arrival Airport'] = iata.text
                alternate_airport = iata.text
                break

    if not alternate_airport:
        fuel_data['Alternate Arrival Airport'] = "N/A"

    alternate_fuels = root.findall('.//ns:AlternateFuel', namespaces)
    if alternate_fuels:
        fuel_value = alternate_fuels[0].find('.//ns:EstimatedWeight/ns:Value', namespaces)
        fuel_data['AlternateFuel'] = float(fuel_value.text) / 1000 if fuel_value is not None and fuel_data['Alternate Arrival Airport'] != 'N/A' and fuel_value.text.replace('.', '', 1).isdigit() else 0.0

    possible_extra = root.find('.//ns:PossibleExtra/ns:EstimatedWeight/ns:Value', namespaces)
    fuel_data['Fuel for other safety rules (tonnes)'] = float(possible_extra.text) / 1000 if possible_extra is not None and possible_extra.text.replace('.', '', 1).isdigit() else 0.0

    fuel_on_board = root.find('.//ns:FuelOnBoard/ns:EstimatedWeight/ns:Value', namespaces)
    fuel_data['FOB'] = float(fuel_on_board.text) / 1000 if fuel_on_board is not None and fuel_on_board.text.replace('.', '', 1).isdigit() else 0.0

    for key, xpath in {
        'DepartureAirport': './/ns:DepartureAirport/ns:AirportICAOCode',
        'ArrivalAirport': './/ns:ArrivalAirport/ns:AirportICAOCode'
    }.items():
        element = root.find(xpath, namespaces)
        fuel_data[key] = element.text if element is not None else ""

    flight = root.find('.//ns:Flight', namespaces)
    flight_number = ""
    if flight is not None:
        flight_number = flight.get('flightNumber')
        if not flight_number:
            flight_number_element = flight.find('.//ns:FlightNumber', namespaces)
            if flight_number_element is not None:
                flight_number = flight_number_element.get('number')

    fuel_data['Flight Number'] = flight_number

    # Extraction CORRIGÉE de la date et heure de départ depuis les attributs de Flight
    flight_element = root.find('.//ns:Flight', namespaces)
    fuel_data['Date of Flight'] = ""
    fuel_data['Time of Departure'] = ""
    
    if flight_element is not None:
        # Extraire la date de vol depuis flightOriginDate
        flight_origin_date = flight_element.get('flightOriginDate')
        if flight_origin_date:
            try:
                # Supprimer le 'Z' final et gérer les formats possibles
                date_clean = flight_origin_date.replace('Z', '').strip()
                # Essayer différents formats de date
                for fmt in ["%Y-%m-%d", "%Y-%m-%d %H:%M:%S"]:
                    try:
                        dt_date = datetime.strptime(date_clean, fmt)
                        fuel_data['Date of Flight'] = dt_date.strftime('%d/%m/%Y')
                        break
                    except ValueError:
                        continue
                else:
                    print(f"Erreur de format de flightOriginDate: {flight_origin_date}")
            except Exception as e:
                print(f"Erreur lors du traitement de flightOriginDate: {flight_origin_date} - {e}")

        # Extraire l'heure de départ depuis scheduledTimeOfDeparture
        scheduled_departure = flight_element.get('scheduledTimeOfDeparture')
        if scheduled_departure:
            try:
                # Supprimer le 'Z' final
                time_clean = scheduled_departure.replace('Z', '').strip()
                # Extraire uniquement la partie heure (après le 'T')
                time_part = time_clean.split('T')[1] if 'T' in time_clean else time_clean
                # Convertir en objet datetime pour formater correctement
                dt_time = datetime.strptime(time_part, "%H:%M:%S")
                fuel_data['Time of Departure'] = dt_time.strftime('%H:%M:%S')
            except (ValueError, IndexError, AttributeError) as e:
                print(f"Erreur de format de scheduledTimeOfDeparture: {scheduled_departure} - {e}")

    # Extraction des raisons de tankering
    tankering_advice = root.find('.//ns:TankeringAdvice', namespaces)
    reason_tankering = tankering_advice.get('reason', "") if tankering_advice is not None else ""

    possible_extra_fuel = root.find('.//ns:PossibleExtraFuel', namespaces)
    reason_possible_extra = possible_extra_fuel.get('reason', "") if possible_extra_fuel is not None else ""

    fuel_data['Reason'] = reason_tankering if reason_tankering else (reason_possible_extra if reason_possible_extra else "")
    fuel_data['Economic tankering category in the flight plan'] = reason_tankering

    if 'Date of Flight' in fuel_data:
        fuel_data['Date of Flight'] = format_date(fuel_data['Date of Flight'])

    discretionary_fuel = fuel_data['BlockFuel'] - fuel_data['FOB']
    fuel_data['Discretionary Fuel'] = round(abs(discretionary_fuel), 3)

    # Extraction de Extra Fuel
    extra_fuel = root.find('.//ns:ExtraFuel/ns:EstimatedWeight/ns:Value', namespaces)
    fuel_data['Extra Fuel'] = float(extra_fuel.text) / 1000 if extra_fuel is not None and extra_fuel.text and extra_fuel.text.replace('.', '', 1).isdigit() else 0.0

    # Extraction de Air Distance
    air_distance = root.find('.//ns:AirDistance/ns:Value', namespaces)
    fuel_data['Air Distance (NM)'] = float(air_distance.text) if air_distance is not None and air_distance.text.replace('.', '', 1).isdigit() else 0.0

    # Calcul des émissions de carbone
    EF = 3.1  # Facteur d'émission
    X = fuel_data['BlockFuel'] + fuel_data['Extra Fuel'] - fuel_data['FOB']
    fuel_data['Carbon Emission (kg)'] = round(X * EF, 3) if X >= 0 else 0.0

    return fuel_data

def process_xml_files(parent_directory):
    if not os.path.isdir(parent_directory):
        print(f"DEBUG: Le répertoire {parent_directory} n'existe pas ou n'est pas un répertoire valide.")
        return json.dumps({'success': False, 'message': f"Erreur : {parent_directory} n'est pas un répertoire valide."})

    all_data = []
    print(f"DEBUG: Début de l'exploration du répertoire {parent_directory}")

    for root_dir, dirs, files in os.walk(parent_directory):
        print(f"DEBUG: Exploration du dossier : {root_dir}")
        print(f"DEBUG: Fichiers trouvés : {files}")
        for filename in files:
            if filename.lower() == "ofp.xml":  # Case-insensitive check
                file_path = os.path.join(root_dir, filename)
                print(f"DEBUG: Traitement du fichier XML : {file_path}")
                try:
                    with open(file_path, 'r', encoding='utf-8') as file:
                        xml_content = file.read()
                    data = extract_fuel_data(xml_content, root_dir)
                    if data:
                        all_data.append(data)
                        print(f"DEBUG: Données extraites avec succès pour {file_path}")
                except Exception as e:
                    print(f"Erreur lors du traitement du fichier {file_path}: {e}")

    if all_data:
        column_order = [
            'Date of Flight', 
            'Time of Departure', 
            'Flight Number', 
            'DepartureAirport', 
            'ArrivalAirport', 
            'TaxiFuel', 
            'TripFuel', 
            'ContingencyFuel', 
            'BlockFuel', 
            'FinalReserve',  
            'Additional Fuel (tonnes)',  
            'Fuel for other safety rules (tonnes)', 
            'Discretionary Fuel', 
            'Extra Fuel',  
            'Reason', 
            'Economic tankering category in the flight plan', 
            'AlternateFuel', 
            'Alternate Arrival Airport',  
            'FOB',
            'Air Distance (NM)',  # Nouvelle colonne
            'Carbon Emission (kg)'  # Nouvelle colonne
        ]
        
        df = pd.DataFrame(all_data)
        for col in column_order:
            if col not in df.columns:
                df[col] = ""
                
        df = df[column_order]
        
        output_excel = os.path.join(parent_directory, "all_fuel_data.xlsx")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Flight Fuel Data"

        for col_num, header in enumerate(column_order, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20

        for row_num, row_data in enumerate(df.to_dict('records'), 2):
            for col_num, col_name in enumerate(column_order, 1):
                value = row_data[col_name]
                if isinstance(value, (int, float)):
                    value = round(value, 3)
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = Alignment(horizontal='right', vertical='center')

        border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border

        try:
            wb.save(output_excel)
            print(f"DEBUG: Fichier Excel sauvegardé : {output_excel}")
            return json.dumps({
                'success': True, 
                'message': f"Téléchargement des données terminé : {output_excel}", 
                'data': all_data
            })
        except PermissionError:
            print(f"DEBUG: ErreurNicolas permission lors de la sauvegarde de {output_excel}")
            return json.dumps({
                'success': False, 
                'message': f"Erreur : Impossible d'enregistrer le fichier. Vérifiez qu'il n'est pas ouvert dans Excel."
            })
    else:
        print(f"DEBUG: Aucun fichier XML valide trouvé dans {parent_directory}")
        return json.dumps({'success': False, 'message': f"Aucun fichier XML trouvé dans {parent_directory}"})

# Définir le répertoire parent
parent_directory = "/app/datax/data"

# Exécuter le script
result = process_xml_files(parent_directory)

# Convertir la chaîne JSON en dictionnaire
result_dict = json.loads(result)

print("\n📜 Résultat brut en JSON :", json.dumps(result_dict, indent=4, ensure_ascii=False))

if result_dict["success"]:
    print("✅ Extraction réussie ! Voici les données extraites :\n")
    for data in result_dict.get("data", []):
        ordered_data = {
            'Date of Flight': data.get('Date of Flight', ''),
            'Time of Departure': data.get('Time of Departure', ''),
            'Flight Number': data.get('Flight Number', ''),
            'DepartureAirport': data.get('DepartureAirport', ''),
            'ArrivalAirport': data.get('ArrivalAirport', ''),
            'TaxiFuel': data.get('Taxi Fuel', 0),
            'TripFuel': data.get('TripFuel', 0),
            'ContingencyFuel': data.get('ContingencyFuel', 0),
            'BlockFuel': data.get('BlockFuel', 0),
            'FinalReserve': data.get('FinalReserve', 0),  
            'Additional Fuel (tonnes)': data.get('Additional Fuel (tonnes)', 0),  
            'Fuel for other safety rules (tonnes)': data.get('Fuel for other safety rules (tonnes)', 0),
            'Discretionary Fuel': data.get('Discretionary Fuel', 0),
            'Extra Fuel': data.get('Extra Fuel', 0),  
            'Reason': data.get('Reason', ''),
            'Economic tankering category in the flight plan': data.get('Economic tankering category in the flight plan', ''),
            'AlternateFuel': data.get('AlternateFuel', 0),
            'Alternate Arrival Airport': data.get('Alternate Arrival Airport', ''),  
            'FOB': data.get('FOB', 0),
            'Air Distance (NM)': data.get('Air Distance (NM)', 0),  # Nouvelle colonne
            'Carbon Emission (kg)': data.get('Carbon Emission (kg)', 0)  # Nouvelle colonne
        }
        print(json.dumps(ordered_data, indent=4, ensure_ascii=False))
else:
    print(f"❌ Erreur : {result_dict['message']}")